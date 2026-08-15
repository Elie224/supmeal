import openpyxl
import shutil
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from collections import Counter

# Reload source files
wb_v3 = openpyxl.load_workbook("user_file_v3.xlsx", data_only=True)
ws_v3 = wb_v3.active

v3_products = {}
for r in range(2, ws_v3.max_row+1):
    cb = ws_v3.cell(r, 2).value
    if cb:
        v3_products[str(cb).replace(" ","")] = {
            "Code EU": ws_v3.cell(r, 1).value or "",
            "Code Barre": str(cb).replace(" ",""),
            "Designation": ws_v3.cell(r, 3).value or "",
            "Longueur UC mm": ws_v3.cell(r, 4).value,
            "Largeur UC mm": ws_v3.cell(r, 5).value,
            "Hauteur UC mm": ws_v3.cell(r, 6).value,
            "Longueur carton mm": ws_v3.cell(r, 7).value,
            "Largeur carton mm": ws_v3.cell(r, 8).value,
            "Hauteur carton mm": ws_v3.cell(r, 9).value,
        }

print(f"Source user v3: {len(v3_products)} produits")

# Load EURODELICES matrix
wb_m = openpyxl.load_workbook("Eurodelices_local.xlsx", data_only=True)
ws_m = wb_m["Matrice de commande"]

matrix_products = {}
for r in range(20, ws_m.max_row+1):
    ean = ws_m.cell(r, 6).value
    if not ean: continue
    ean_clean = str(ean).replace(" ","")
    if ean_clean in matrix_products: continue  # dedupe
    matrix_products[ean_clean] = {
        "SKU": str(ws_m.cell(r, 3).value or ""),
        "Designation": str(ws_m.cell(r, 5).value or ""),
        "Categorie": str(ws_m.cell(r, 16).value or ""),
    }

print(f"Matrice MORMANT: {len(matrix_products)} produits uniques")

# Combine: products from v3 take priority
final = {}
# First pass: v3 (with dimensions)
for ean, info in v3_products.items():
    final[ean] = info.copy()

# Second pass: matrix MORMANT (additions only)
added_from_matrix = 0
for ean, info in matrix_products.items():
    if ean not in final:
        final[ean] = {
            "Code EU": "",
            "Code Barre": ean,
            "Designation": info["Designation"],
            "Longueur UC mm": "",
            "Largeur UC mm": "",
            "Hauteur UC mm": "",
            "Longueur carton mm": "",
            "Largeur carton mm": "",
            "Hauteur carton mm": "",
        }
        added_from_matrix += 1

print(f"Depuis MORMANT (ajoutes): {added_from_matrix}")
print(f"Total final (sans dedup intra-MORMANT): {len(final)}")

# Check duplicates within v3
v3_eans = [str(ws_v3.cell(r,2).value).replace(" ","") for r in range(2, ws_v3.max_row+1) if ws_v3.cell(r,2).value]
dup_v3 = [e for e, n in Counter(v3_eans).items() if n > 1]
print(f"\nDuplicats dans v3: {len(dup_v3)}")
for e in dup_v3:
    print(f"  {e}")

# Write clean output
output = "MATRICE_MONDELEZ_FINALE.xlsx"
wb_out = openpyxl.Workbook()
ws_out = wb_out.active
ws_out.title = "Matrice logistique"

headers = ["Code EU", "Code Barre", "Designation",
           "Longueur UC mm", "Largeur UC mm", "Hauteur UC mm",
           "Longueur carton mm", "Largeur carton mm", "Hauteur carton mm"]

header_font = Font(bold=True, color="FFFFFF", size=11)
header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
center = Alignment(horizontal="center", vertical="center", wrap_text=True)
border = Border(left=Side(style="thin"), right=Side(style="thin"),
                top=Side(style="thin"), bottom=Side(style="thin"))

for col_idx, h in enumerate(headers, 1):
    c = ws_out.cell(1, col_idx, h)
    c.font = header_font
    c.fill = header_fill
    c.alignment = center
    c.border = border

# Sort: by Code EU, then Designation
sorted_p = sorted(final.values(), key=lambda p: (p["Code EU"] or "ZZZZZ", p["Designation"]))

for row_idx, p in enumerate(sorted_p, 2):
    for col_idx, h in enumerate(headers, 1):
        v = p.get(h, "")
        if v is None: v = ""
        c = ws_out.cell(row_idx, col_idx, v)
        c.alignment = center
        c.border = border

widths = {"A": 12, "B": 16, "C": 55, "D": 14, "E": 14, "F": 14, "G": 16, "H": 16, "I": 16}
for col, w in widths.items():
    ws_out.column_dimensions[col].width = w

ws_out.row_dimensions[1].height = 35
ws_out.freeze_panes = "A2"

wb_out.save(output)
print(f"\nFichier sauvegarde: {output}")

total = ws_out.max_row - 1
with_eu = sum(1 for p in final.values() if p["Code EU"])
with_dims = sum(1 for p in final.values() if p["Longueur UC mm"])
print(f"Total: {total}")
print(f"Avec Code EU: {with_eu}")
print(f"Avec dimensions: {with_dims}")
print(f"Sans dimensions: {total - with_dims}")
