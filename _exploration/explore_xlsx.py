import pandas as pd
import openpyxl

files = {
    '2024': r'C:\\Users\\KOURO\\Downloads\\RECAP LAITS DECLASSES 2024 (1) (1).xlsx',
    '2025': r'C:\\Users\\KOURO\\Downloads\\RECAP LAITS DECLASSES 2025 (1) (1).xlsx',
    '25_26': r'C:\\Users\\KOURO\\Downloads\\RECAP LAITS DECLASSES 25 26.xlsx',
}

for label, f in files.items():
    print('=' * 80)
    print(f'FICHIER: {label} -> {f}')
    print('=' * 80)
    wb = openpyxl.load_workbook(f, data_only=False)
    for sh in wb.sheetnames:
        ws = wb[sh]
        print(f'  Sheet: {sh!r}  dim={ws.dimensions}  rows={ws.max_row} cols={ws.max_column}')
    print()
